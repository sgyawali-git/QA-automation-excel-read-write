import os
import openpyxl
#full location of the folder with excel name where new excel file get created
#used for test result

test_result_location = r"C:\Users\A C E R\PycharmProjects\Python Learning\swag_lab\testResult/test_result.xlsx"
def create_excel():
    wb = openpyxl.Workbook()
    ws = wb.create_sheet("result")

    ws.cell(1,1,"SN")
    ws.cell(1, 2, "Test Summary")
    ws.cell(1, 3, "Result")
    ws.cell(1, 4, "Remarks")
    wb.save(test_result_location)

def remove_file():
    #to remove created files before creating new excel files
        os.remove(r"C:\Users\A C E R\PycharmProjects\Python Learning\swag_lab\testResult\test_result.xlsx")
def write_result(sn,test_summary,result,remarks):
    wb = openpyxl.load_workbook(test_result_location)
    ws = wb["result"]
    row = int(sn)+1
    ws.cell(row,1,sn)
    ws.cell(row, 2, test_summary)
    ws.cell(row, 3, result)
    ws.cell(row, 4, str(remarks))
    wb.save(test_result_location)


